import re
import ast
import sys
import tkinter as tk
from tkinter import simpledialog
from .base_operator import BaseOperator
from openpyxl.utils import column_index_from_string, get_column_letter


class TableOperator(BaseOperator):
    """
    Оператор TABLE принимает два и более списков в квадратных скобках.
    Первый список задаёт заголовки столбцов, последующие — строки данных.
    Элемент 'INPUT' создаёт поле ввода. Элемент, начинающийся с '=', трактуется как формула,
    поддерживающая ссылки на ячейки (A1, B2 и т.д.) и переменные из variable_manager.
    Отображает окно Tkinter с таблицей и кнопками "Продолжить" и "Остановиться".
    """

    def execute(self, command):
        # Парсим списки из команды
        lists = re.findall(r"\[[^\]]*\]", command)
        if len(lists) < 2:
            raise ValueError("Оператор TABLE требует минимум два списка в квадратных скобках")
        try:
            data = [ast.literal_eval(lst) for lst in lists]
        except Exception as e:
            raise ValueError(f"Ошибка разбора списков: {e}")
        headers = data[0]
        rows = data[1:]

        # Создаём окно
        root = tk.Tk()
        root.title("Input Table")
        entries = {}
        formula_labels = {}  # Для хранения Label с результатами формул

        # Рисуем заголовки
        for j, header in enumerate(headers):
            lbl = tk.Label(root, text=str(header), font=('Arial', 12, 'bold'),
                           borderwidth=1, relief="solid", padx=5, pady=5)
            lbl.grid(row=0, column=j, sticky="nsew")

        # Рисуем ячейки
        for i, row in enumerate(rows, start=1):
            for j, val in enumerate(row):
                cell_id = f"{get_column_letter(j + 1)}{i + 1}"
                if isinstance(val, str) and val.upper() == 'INPUT':
                    ent = tk.Entry(root)
                    ent.grid(row=i, column=j, sticky="nsew")
                    entries[cell_id] = ent
                elif isinstance(val, str) and val.startswith('='):
                    # Создаём Label для отображения результата формулы
                    lbl = tk.Label(root, text="", borderwidth=1, relief="solid",
                                   padx=5, pady=5)
                    lbl.grid(row=i, column=j, sticky="nsew")
                    formula_labels[cell_id] = (lbl, val)  # Сохраняем виджет и формулу
                else:
                    lbl = tk.Label(root, text=str(val), borderwidth=1, relief="solid",
                                   padx=5, pady=5)
                    lbl.grid(row=i, column=j, sticky="nsew")

        # Кнопки
        result = {'continue': False}

        def on_continue():
            result['continue'] = True
            root.quit()

        def on_stop():
            root.destroy()
            sys.exit(0)

        btn_continue = tk.Button(root, text="Продолжить", command=on_continue, padx=10, pady=5)
        btn_stop = tk.Button(root, text="Остановиться", command=on_stop, padx=10, pady=5)
        btn_continue.grid(row=len(rows) + 1, column=0, columnspan=len(headers) // 2, pady=10)
        btn_stop.grid(row=len(rows) + 1, column=len(headers) // 2, columnspan=len(headers) - len(headers) // 2, pady=10)

        # Автоматическая растяжка колонок
        for c in range(len(headers)):
            root.grid_columnconfigure(c, weight=1)

        # Функция для пересчета формул
        def recalculate_formulas():
            # Собираем введенные значения
            collected = {}
            for cell_id, ent in entries.items():
                collected[cell_id] = ent.get()

            # Создаем словарь значений всех ячеек
            cells = {}
            for i, row in enumerate(rows, start=1):
                for j, val in enumerate(row):
                    cell_id = f"{get_column_letter(j + 1)}{i + 1}"
                    if isinstance(val, str) and val.upper() == 'INPUT':
                        cells[cell_id] = collected.get(cell_id, '0')
                    else:
                        cells[cell_id] = val

            # Функция для обработки диапазонов
            def _range(range_str):
                parts = [p.strip() for p in range_str.split(':', 1)]
                if len(parts) != 2:
                    return []
                start, end = parts
                start_col = re.match(r'[A-Z]+', start).group()
                start_row = int(re.search(r'\d+$', start).group())
                end_col = re.match(r'[A-Z]+', end).group()
                end_row = int(re.search(r'\d+$', end).group())

                start_col_idx = column_index_from_string(start_col)
                end_col_idx = column_index_from_string(end_col)

                values = []
                for row_num in range(start_row, end_row + 1):
                    for col_idx in range(start_col_idx, end_col_idx + 1):
                        col_letter = get_column_letter(col_idx)
                        cell_id = f"{col_letter}{row_num}"
                        value = cells.get(cell_id, '0')
                        try:
                            num = float(value)
                            values.append(num)
                        except (ValueError, TypeError):
                            values.append(0)
                return values

            # Пользовательские функции
            def average(*args):
                total = 0.0
                count = 0
                for arg in args:
                    if isinstance(arg, list):
                        for x in arg:
                            total += x
                            count += 1
                    else:
                        total += arg
                        count += 1
                return total / count if count != 0 else 0

            def sum_excel(*args):
                total = 0.0
                for arg in args:
                    if isinstance(arg, list):
                        total += sum(arg)
                    else:
                        total += arg
                return total

            def min_excel(*args):
                flat_vals = []
                for arg in args:
                    if isinstance(arg, list):
                        flat_vals.extend(arg)
                    else:
                        flat_vals.append(arg)
                return min(flat_vals) if flat_vals else 0

            def max_excel(*args):
                flat_vals = []
                for arg in args:
                    if isinstance(arg, list):
                        flat_vals.extend(arg)
                    else:
                        flat_vals.append(arg)
                return max(flat_vals) if flat_vals else 0

            # Сопоставление русских и английских названий функций
            function_map = {
                "СРЗНАЧ": "average",
                "СУММ": "sum_excel",
                "SUM": "sum_excel",
                "AVERAGE": "average",
                "MIN": "min_excel",
                "MAX": "max_excel",
                "ABS": "abs",
            }

            # Контекст для вычисления формул
            context = {
                '_range': _range,
                'average': average,
                'sum_excel': sum_excel,
                'min_excel': min_excel,
                'max_excel': max_excel,
                'abs': abs,
            }
            context.update(self.vm.get_all_variables())

            # Вычисляем формулы и обновляем Label
            for cell_id, (label_widget, formula) in formula_labels.items():
                expr = formula.lstrip('=')

                # Замена диапазонов
                expr = re.sub(r'([A-Z]+\d+\s*:\s*[A-Z]+\d+)', r'_range("\1")', expr)

                # Замена названий функций
                for ru_name, en_name in function_map.items():
                    expr = expr.replace(ru_name, en_name)

                # Замена одиночных ячеек
                def repl(match):
                    ref = match.group(0)
                    value = cells.get(ref, '0')
                    try:
                        return str(float(value))
                    except (ValueError, TypeError):
                        return '0'

                expr = re.sub(r'[A-Z]+\d+', repl, expr)

                # Вычисление выражения
                try:
                    result = eval(expr, {}, context)
                    if isinstance(result, float):
                        # Форматируем float для отображения
                        label_widget.config(text=f"{result:.2f}")
                    else:
                        label_widget.config(text=str(result))
                except Exception as e:
                    label_widget.config(text=f"#ERR: {e}")

        # Первоначальный расчет формул
        recalculate_formulas()

        # Привязываем пересчет формул к изменению полей ввода
        def on_input_change(event):
            recalculate_formulas()

        for ent in entries.values():
            ent.bind('<KeyRelease>', on_input_change)

        root.mainloop()
        root.destroy()

        # Если пользователь нажал "Остановиться", программа завершится в on_stop
        if not result['continue']:
            return {}

        # Собираем финальные данные
        collected = {}
        for cell_id, ent in entries.items():
            collected[cell_id] = ent.get()

        # Создаем словарь всех ячеек
        cells = {}
        for i, row in enumerate(rows, start=1):
            for j, val in enumerate(row):
                cell_id = f"{get_column_letter(j + 1)}{i + 1}"
                if cell_id in formula_labels:
                    # Берем результат из Label
                    cells[cell_id] = formula_labels[cell_id][0].cget("text")
                elif isinstance(val, str) and val.upper() == 'INPUT':
                    cells[cell_id] = collected.get(cell_id, '')
                else:
                    cells[cell_id] = val

        return cells