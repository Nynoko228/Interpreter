import re
import ast
import sys
import tkinter as tk
from tkinter import messagebox
from .base_operator import BaseOperator
from openpyxl.utils import get_column_letter, column_index_from_string
from core.protocol_utils.protocol_utils import get_base_cells_in_range

class WriteOperator(BaseOperator):
    def execute(self, command):
        # Удаляем ключевое слово WRITE
        expr = re.sub(r'^(WRITE|ЗАПИСЬ)\s*', '', command, flags=re.IGNORECASE).strip()

        # Парсинг источника и цели
        if '->' not in expr:
            raise ValueError("Неверный формат команды WRITE. Используйте: WRITE <значение> -> <цель>")

        value_part, target_part = expr.split('->', 1)
        value_part = value_part.strip()
        target_part = target_part.strip()

        # Получение значения
        value = self._parse_value(value_part)
        # print(f"value в write_operator: {value}")
        # print(type(value))

        # Определение целей
        # print(target_part)
        # print(type(target_part))
        try:
            targets = self._parse_targets(target_part)
        except Exception as e:
            root = tk.Tk()
            root.withdraw()
            messagebox.showerror("Ошибка", f"Не удалось получить метки в протоколе, возможно вы не открыли его командой 'ЗАГРУЗИТЬ_ПРОТОКОЛ':\n{e}")
            root.destroy()
            # Завершаем программу с кодом ошибки
            sys.exit(1)
        # print(f"targets в write_operator: {targets}")
        # print(type(targets))
        # print(type(value), type(targets))
        # if len(value) != len(targets):
        #     raise Exception("Количество элементов для записи и количество ячеек должны быть равны")
        # Запись в протокол
        self._write_to_protocol(targets, value)

        # Логирование
        # print(f"WRITE: {value} -> {targets}")
        return value

    def _parse_value(self, value_part):
        """Анализирует и преобразует значение"""
        # Попытка извлечь значение из переменной
        try:
            return self.vm.get_variable(value_part)[1]
        except:
            pass
        # Строковые значения
        if (value_part.startswith('"') and value_part.endswith('"')) or \
                (value_part.startswith("'") and value_part.endswith("'")):
            value_part = super()._replace_variables(value_part[1:-1])
            return value_part

        # Булевы значения
        if value_part.lower() in ('true', 'истина'):
            return True
        if value_part.lower() in ('false', 'ложь'):
            return False

        # Списки значений
        if value_part.startswith('[') and value_part.endswith(']'):
            try:
                return ast.literal_eval(value_part)
            except (ValueError, SyntaxError):
                # Обработка простых списков без кавычек
                items = [item.strip() for item in value_part[1:-1].split(',')]
                return [self._parse_value(item) for item in items]

        # Числовые значения
        try:
            return float(value_part)
        except ValueError:
            pass

        # Попытка как выражение
        try:
            return self.vm.evaluate_expression(value_part)
        except:
            pass

        # По умолчанию возвращаем как строку
        return value_part

    def _resolve_tag_to_cell(self, tag):
        """Преобразует метку в адрес ячейки"""
        clean_tag = tag.strip('{}').strip()
        try:
            # Получаем координаты метки из протокола
            tags_dict = self.vm.protocol.tags.get(0, {})
            print(f"tags_dict: {tags_dict}")
            if clean_tag in tags_dict:
                row, col = tags_dict[clean_tag]
                return f"{get_column_letter(col)}{row}"
        except Exception as e:
            print(f"Ошибка при разрешении метки {tag}: {str(e)}")
        return None

    def _parse_targets(self, target_part):
        """Анализирует целевые спецификации"""
        tp = target_part.strip()
        # Список целей в квадратных скобках, с поддержкой диапазонов
        if tp.startswith('[') and tp.endswith(']'):
            content = tp[1:-1]
            items = [item.strip() for item in content.split(',')]
            results = []
            for item in items:
                if ':' in item:
                    start, end = [x.strip() for x in item.split(':', 1)]
                    results.extend(self._expand_cell_range(start, end))
                else:
                    results.append(self._normalize_target(item))
            return results
        # Простой диапазон без скобок
        if ':' in tp:
            start, end = [x.strip() for x in tp.split(':', 1)]
            return self._expand_cell_range(start, end)
        # Одиночная цель
        return [self._normalize_target(tp)]


    # def _parse_targets(self, target_part):
    #     """Анализирует целевые спецификации"""
    #     # print(f"_parse_targets: {target_part}")
    #     # print(':' in target_part)
    #
    #     # Диапазон ячеек A1:B3
    #     if ':' in target_part:
    #         start, end = target_part.split(':', 1)
    #         start = start.strip()
    #         end = end.strip()
    #
    #         # Преобразуем в координаты
    #         start_coords = self._parse_target_to_coords(start)
    #         end_coords = self._parse_target_to_coords(end)
    #         # print(start_coords, end_coords)
    #         # print(start, end)
    #         if not start_coords or not end_coords:
    #             raise ValueError(f"Неверный формат диапазона: {start}:{end}")
    #
    #         # Получаем главные ячейки в диапазоне
    #         min_row = min(start_coords[0], end_coords[0])
    #         max_row = max(start_coords[0], end_coords[0])
    #         min_col = min(start_coords[1], end_coords[1])
    #         max_col = max(start_coords[1], end_coords[1])
    #
    #         # Получаем главные ячейки
    #         base_cells = get_base_cells_in_range(
    #             self.vm.protocol.worksheet,
    #             min_row, min_col, max_row, max_col
    #         )
    #
    #         # Преобразуем координаты в адреса
    #         addresses = [f"{get_column_letter(col)}{row}" for row, col in base_cells]
    #         return addresses
    #
    #     # Одиночная цель
    #     if not (target_part.startswith('[') and target_part.endswith(']')):
    #         return [self._normalize_target(target_part)]
    #
    #     # Список целей или диапазон
    #     # print(target_part)
    #     try:
    #         targets = target_part[1:-1].split(',')
    #         # print(f"targets: {targets}")
    #         if isinstance(targets, list):
    #             return [self._normalize_target(t) for t in targets]
    #     except (ValueError, SyntaxError):
    #         pass
    #
    #     # Одиночная цель в скобках
    #     return [self._normalize_target(target_part[1:-1])]

    def _normalize_target(self, target):
        """Нормализует формат цели"""
        # Метка в фигурных скобках
        target = target.strip()
        if target.startswith('{{') and target.endswith('}}'):
            # print(f"Таргет {target} - метка")
            # return target[2:-2].strip()
            return target

        # Адрес ячейки (приводим к верхнему регистру)
        if re.match(r'^[a-zA-Z]+\d+$', target):
            # print(f"Таргет {target} - ячейка")
            return target.upper()

        # print(f"Таргет и тип {target, type(target)}")
        return target

    def _expand_cell_range(self, start, end):
        """Преобразует диапазон ячеек в список главных адресов"""
        # Проверяем, являются ли start/end метками
        if start.startswith('{{') or end.startswith('{{'):
            return self._expand_tag_range(start, end)

        # Извлекаем компоненты
        start_col = ''.join(filter(str.isalpha, start))
        start_row = int(''.join(filter(str.isdigit, start)))
        end_col = ''.join(filter(str.isalpha, end))
        end_row = int(''.join(filter(str.isdigit, end)))

        # Конвертируем в индексы
        col_start = column_index_from_string(start_col)
        col_end = column_index_from_string(end_col)

        # Получаем главные ячейки в диапазоне
        base_cells = get_base_cells_in_range(
            self.vm.protocol.worksheet,
            min(start_row, end_row),
            min(col_start, col_end),
            max(start_row, end_row),
            max(col_start, col_end)
        )

        # Преобразуем координаты в адреса
        addresses = []
        for row, col in base_cells:
            addresses.append(f"{get_column_letter(col)}{row}")
        return addresses

    def _expand_tag_range(self, start_tag, end_tag):
        """Генерирует диапазон между двумя метками, возвращая только главные ячейки"""
        start_coords = self.vm.protocol.get_tag_coordinates(start_tag)
        end_coords = self.vm.protocol.get_tag_coordinates(end_tag)

        if not start_coords or not end_coords:
            raise ValueError(f"Метки диапазона не найдены: {start_tag}:{end_tag}")

        start_row, start_col = start_coords
        end_row, end_col = end_coords

        # Получаем главные ячейки в диапазоне
        base_cells = get_base_cells_in_range(
            self.vm.protocol.worksheet,
            min(start_row, end_row),
            min(start_col, end_col),
            max(start_row, end_row),
            max(start_col, end_col)
        )

        # Преобразуем координаты в адреса ячеек
        addresses = []
        for row, col in base_cells:
            addresses.append(f"{get_column_letter(col)}{row}")
        return addresses

    def _write_to_protocol(self, targets, value):
        """Записывает значение в указанные цели протокола"""
        # Для одиночной цели
        if len(targets) == 1 and len([value]) == 1:
            self._write_single_target(targets[0], value)
            return

        # Для нескольких целей
        if not isinstance(value, (list, tuple)):
            # print("Несколько targets")
            # Одно значение для всех целей
            for target in targets:
                self._write_single_target(target, value)
        else:
            # Проверка соответствия количества
            # if len(value) != len(targets):
            #     raise ValueError(
            #         f"Количество значений ({len(value)}) должно быть равным 1 или совпадать с количеством ячеек таблицы ({len(targets)})")

            # Распределение значений по целям
            # for i in range(min(len(targets), len(value))):
            for i in range(len(value)):
                self._write_single_target(targets[i], value[i])

    def _parse_target_to_coords(self, target_str):
        """Преобразует цель в координаты (строка, столбец)"""
        # Обработка меток
        if target_str.startswith('{{') and target_str.endswith('}}'):
            clean_tag = target_str[2:-2].strip()
            return self._resolve_tag_to_coords(clean_tag)

        # Обработка обычных адресов ячеек
        if re.match(r'^[a-zA-Z]+\d+$', target_str):
            col_letter = ''.join(filter(str.isalpha, target_str))
            row_num = int(''.join(filter(str.isdigit, target_str)))
            col_num = column_index_from_string(col_letter)
            return (row_num, col_num)

        # Попытка обработки как метки без скобок
        return self._resolve_tag_to_coords(target_str)

    def _resolve_tag_to_coords(self, tag):
        """Преобразует метку в координаты (строка, столбец)"""
        clean_tag = tag.strip()
        try:
            # Получаем координаты метки из протокола
            # tags_dict = self.vm.protocol.tags.items()
            print(f"tags_dict: {self.vm.protocol.tags}")
            print(tag)
            for i in self.vm.protocol.tags:
                for j in self.vm.protocol.tags[i]:
                    if tag == j:
                        return self.vm.protocol.tags[i][j]
                    # print(self.vm.protocol.tags[i][j] if tag == j else False)
                # print(self.vm.protocol.tags[i])
            # print(tags_dict['Лист1'])
            # for i in self.vm.protocol.tags.keys():
            #     if clean_tag in tags_dict[i]:
            #
            #         return tags_dict[clean_tag]
        except Exception as e:
            # print(e)
            print(f"Ошибка при разрешении метки {tag}: {str(e)}")
        return None

    def _write_single_target(self, target, value):
        """Записывает значение в одну цель"""
        protocol = self.vm.protocol
        # Определяем тип цели
        if self._is_cell_reference(target):
            # Запись по адресу ячейки
            protocol.add_cell_replacement(target, value)
        else:
            # Запись по метке
            protocol.add_tag_replacement(target, value)

    def _is_cell_reference(self, ref):
        """Проверяет, является ли цель ссылкой на ячейку"""
        return re.match(r'^[A-Z]+\d+$', ref) is not None

