from io import BytesIO

import openpyxl
import re


def find_all_tags(file_path):
    wb = openpyxl.load_workbook(BytesIO(file_path))
    results = {}

    # Регулярное выражение для поиска меток вида {{...}}
    pattern = re.compile(r'\{\{(.*?)\}\}')

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_tags = {}

        # Собираем объединенные ячейки для исключения дубликатов
        merged_cells = set()
        for merge_range in ws.merged_cells.ranges:
            min_row, min_col, max_row, max_col = merge_range.min_row, merge_range.min_col, merge_range.max_row, merge_range.max_col
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    if row != min_row or col != min_col:
                        merged_cells.add((row, col))

        # Обрабатываем все ячейки в листе
        for row in ws.iter_rows():
            for cell in row:
                # Пропускаем ячейки в объединенных областях
                if (cell.row, cell.column) in merged_cells:
                    continue

                # Ищем метки в значении ячейки
                if cell.value and isinstance(cell.value, str):
                    matches = pattern.findall(cell.value)
                    for tag in matches:
                        clean_tag = tag.strip()
                        sheet_tags[clean_tag] = [cell.row, cell.column]

        if sheet_tags:
            results[sheet_name] = sheet_tags

    return results


def get_merged_cell_mapping(ws):
    """Создает отображение для объединенных ячеек (ячейка -> главная ячейка)"""
    mapping = {}
    for merged_range in ws.merged_cells.ranges:
        min_row, min_col = merged_range.min_row, merged_range.min_col
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                mapping[(row, col)] = (min_row, min_col)
    return mapping


def get_base_cells_in_range(ws, min_row, min_col, max_row, max_col):
    """Возвращает уникальные главные ячейки в указанном диапазоне"""
    merged_mapping = get_merged_cell_mapping(ws)
    base_cells = set()

    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            base_cell = merged_mapping.get((row, col), (row, col))
            base_cells.add(base_cell)

    return sorted(base_cells)  # Сортируем для сохранения порядка

if __name__ == "__main__":
    print(find_all_tags("C:\MeMorozov\Interpreter\Protocol poverki\Testiruem\Копия_Термометр_стеклянный_(1).xlsx"))