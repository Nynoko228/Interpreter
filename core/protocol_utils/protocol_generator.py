import sys

from openpyxl import load_workbook
import re
import tkinter as tk
from tkinter import messagebox
from collections import defaultdict
from openpyxl.utils import get_column_letter


class ProtocolGenerator:
    def __init__(self, template_path=None):
        self.template_path = template_path
        self.cell_replacements = {}
        self.tag_replacements = defaultdict(list)
        self.tags = {}

    def set_template_path(self, path):
        self.template_path = path

    def add_cell_replacement(self, cell_ref, value):
        """Добавляет замену для ячейки по координатам"""
        self.cell_replacements[cell_ref] = value

    def add_tag_replacement(self, tag, value):
        """Добавляет замену для метки"""
        # Удаляем фигурные скобки если есть
        clean_tag = tag.strip('{}').strip()
        self.tag_replacements[clean_tag].append(value)

    def set_tags(self, tags):
        """Устанавливает словарь меток (метка -> [ячейки])"""
        self.tags = tags

    # Функция для поиска главной ячейки в объединённом диапазоне
    def get_base_cell(self, ws, row, col):
        """Возвращает координаты главной ячейки объединённого диапазона"""
        for merged_range in ws.merged_cells.ranges:
            if (merged_range.min_row <= row <= merged_range.max_row and
                    merged_range.min_col <= col <= merged_range.max_col):
                return merged_range.min_row, merged_range.min_col
        return row, col

    def generate(self, output_path):
        """Генерирует протокол с учетом всех замен"""
        if not self.template_path:
            raise ValueError("Путь к шаблону не установлен")

        wb = load_workbook(self.template_path)
        ws = wb.active
        print("GENERATE")
        # Применяем замены по ячейкам
        for cell_ref, value in self.cell_replacements.items():
            try:
                col_letter = ''.join(filter(str.isalpha, cell_ref))
                row_num = int(''.join(filter(str.isdigit, cell_ref)))
                col_index = self._column_index(col_letter)

                # Определяем главную ячейку объединения
                target_row, target_col = self.get_base_cell(ws, row_num, col_index)

                # Проверяем, является ли ячейка объединённой
                if (target_row, target_col) != (row_num, col_index):
                    print(f"Внимание: Ячейка {cell_ref} находится в объединённом диапазоне. "
                          f"Запись будет произведена в главную ячейку {get_column_letter(target_col)}{target_row}")

                ws.cell(row=target_row, column=target_col).value = value

            except Exception as e:
                print(f"Ошибка при обработке ячейки {get_column_letter(target_col)}{target_row}")
                print(f"Значение '{value}' не было записано в ячейку {cell_ref}")

        # Применяем замены по меткам
        for tag, values in self.tag_replacements.items():
            for i in self.tags:
                if tag in self.tags[i]:
                    try:
                        row = self.tags[i][tag][0]
                        col = self.tags[i][tag][1]

                        # Определяем главную ячейку объединения
                        target_row, target_col = self.get_base_cell(ws, row, col)

                        # Проверяем, является ли ячейка объединённой
                        if (target_row, target_col) != (row, col):
                            print(f"Внимание: Метка '{tag}' [{row}, {col}] находится в объединённом диапазоне. "
                                  f"Запись будет произведена в главную ячейку {get_column_letter(col)}{row}")

                        ws.cell(row=target_row, column=target_col).value = str(values[0])

                    except Exception as e:
                        print(f"Ошибка при обработке метки '{tag}' в ячейку {get_column_letter(col)}{row}: {str(e)}")
                        print(f"Значение '{values[0]}' не было записано для метки '{tag}'")

        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and isinstance(cell.value, str):
                    # Удаляем все вхождения {{...}} из текста ячейки
                    cell.value = re.sub(r'\{\{[^}]*\}\}', '', cell.value)

        try:
            ws.protection.sheet = True
            ws.protection.password = "my_password"
            wb.save(output_path)
            root = tk.Tk()
            root.withdraw()
            messagebox.showinfo("Успех", f"Протокол успешно сохранён:\n{output_path}")
            root.destroy()
            return output_path
        except Exception as e:
            root = tk.Tk()
            root.withdraw()
            messagebox.showerror("Ошибка", f"Не удалось сохранить протокол:\n{e}")
            root.destroy()
            # Завершаем программу с кодом ошибки
            sys.exit(1)

    def _column_index(self, col_letter):
        """Конвертирует букву столбца в номер"""
        index = 0
        for char in col_letter:
            index = index * 26 + (ord(char.upper()) - ord('A') + 1)
        return index

    def get_tag_coordinates(self, tag):
        """Возвращает координаты ячейки по метке"""
        clean_tag = tag.strip('{}').strip()
        for sheet_index, tags_dict in self.tags.items():
            if clean_tag in tags_dict:
                row, col = tags_dict[clean_tag]
                return row, col
        return None
