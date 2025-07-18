import os
import re
import tempfile
from collections import defaultdict

from openpyxl import load_workbook
from io import BytesIO
from datetime import datetime

from core.protocol_utils.protocol_generator import ProtocolGenerator

# Класс для управления переменными
class VariableManager:
    def __init__(self):
        self.variables = {
            'MEM': ('number', 0),
            'MEM1': ('number', 0),
            'MEM2': ('string', ''),
            'HEADER_TOP': ('string', ''),  # Заголовок верхней части
            'HEADER_BOTTOM': ('string', ''),  # Заголовок нижней части
            'RESULTS': ('dict', {}),
        }
        self.protocol_tags = {}
        self.original_protocol_path = None
        self.protocol_data = None  # Бинарные данные протокола
        self.protocol = ProtocolGenerator()
        # self.replacements = defaultdict(dict)  # {sheet_name: {cell_ref: value}}
        # self.protocol_replacements = {
        #     'cells': {},  # {cell_ref: value}
        #     'tags': {}  # {tag: value}
        # }

    def get_variable(self, name):
        var_name = name # Сделаем регистрозависимым
        # var_name = name.upper()
        if var_name not in self.variables:
            raise NameError(f"Переменная {var_name} не определена")
        return self.variables[var_name]

    def set_variable(self, name, var_type, value):
        var_name = name # Сделаем регистрозависимым
        # var_name = name.upper()

        if not re.match(r'^[a-zA-Zа-яА-Я_][a-zA-Z0-9а-яА-Я_]*$', var_name):
            raise NameError(f"Некорректное имя переменной: {var_name}")

        if var_name in self.variables:
            existing_type = self.variables[var_name][0]
            if existing_type != var_type:
                raise TypeError(f"Нельзя изменить тип переменной {var_name}")

        self.variables[var_name] = (var_type, value)
        # print(self.variables)

    def get_all_variables(self):
        return {k: v[1] for k, v in self.variables.items()}

    def set_header_top(self, value):
        """Устанавливает текст в верхнюю часть заголовков"""
        self.set_variable('HEADER_TOP', 'string', value)

    def set_header_bottom(self, value):
        """Устанавливает текст в нижнюю часть заголовков"""
        self.set_variable('HEADER_BOTTOM', 'string', value)

    def get_header_top(self):
        """Получает текст из верхней части заголовков"""
        return self.get_variable('HEADER_TOP')[1]

    def get_header_bottom(self):
        """Получает текст из нижней части заголовков"""
        return self.get_variable('HEADER_BOTTOM')[1]

    def set_protocol_tags(self, tags):
        self.protocol_tags = tags

    def get_protocol_tags(self):
        return self.protocol_tags

    def set_original_protocol_path(self, path):
        self.original_protocol_path = path

    def get_original_protocol_path(self):
        return self.original_protocol_path

    def set_protocol_data(self, data):
        self.protocol_data = data

    def get_protocol_data(self):
        return self.protocol_data

    # def generate_protocol(self, output_path=None):
    #     """Генерирует финальную версию протокола"""
    #     if not self.protocol_data:
    #         raise ValueError("Протокол не загружен")
    #
    #     # Если путь не указан, используем оригинальный путь с отметкой времени
    #     if not output_path:
    #         filename = os.path.basename(self.original_protocol_path)
    #         output_path = os.path.join(
    #             os.path.dirname(self.original_protocol_path),
    #             f"result_{filename}"
    #         )
    #
    #     # Создаем рабочую книгу из бинарных данных
    #     workbook = load_workbook(BytesIO(self.protocol_data))
    #
    #     # Применяем все изменения
    #     for sheet_name, replacements in self.replacements.items():
    #         if sheet_name in workbook.sheetnames:
    #             ws = workbook[sheet_name]
    #             for cell_ref, value in replacements.items():
    #                 ws[cell_ref] = value
    #
    #     # Сохраняем результат
    #     workbook.save(output_path)
    #     return output_path