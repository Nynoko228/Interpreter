from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.cell import MergedCell

def replace_template_values(file_path, output_path, replacements):
    # Загрузка рабочей книги
    wb = load_workbook(filename=file_path)
    ws = wb.active

    # Карта объединённых ячеек: {адрес -> главная ячейка}
    merged_map = {}

    # Строим карту: все ячейки в диапазоне объединения указывают на главную (верхнюю левую)
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = (
            merged_range.min_col, merged_range.min_row,
            merged_range.max_col, merged_range.max_row
        )
        top_left = f"{get_column_letter(min_col)}{min_row}"

        # Обходим все ячейки объединённого диапазона
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                addr = f"{get_column_letter(col)}{row}"
                merged_map[addr] = top_left  # Каждая ячейка диапазона ссылается на главную

    # Применяем замены
    for addr, value in replacements.items():
        col_letter = ''.join(filter(str.isalpha, addr))
        row = int(''.join(filter(str.isdigit, addr)))

        # Проверяем, входит ли ячейка в объединённый диапазон
        if addr in merged_map:
            target_addr = merged_map[addr]
            col_letter = ''.join(filter(str.isalpha, target_addr))
            row = int(''.join(filter(str.isdigit, target_addr)))

        # Преобразуем букву столбца в номер
        col_number = column_index_from_string(col_letter)

        # Получаем ячейку
        cell = ws.cell(row=row, column=col_number)

        # Проверяем, не является ли ячейка частью объединённой (кроме главной)
        if not isinstance(cell, MergedCell):
            print(f"Запись в {addr}: {value}")
            cell.value = value
        else:
            print(f"Пропущено: {addr} (часть объединённой ячейки)")

    # Сохранение изменений
    wb.save(output_path)
    print(f"\nФайл успешно сохранён: {output_path}")


if __name__ == "__main__":
    # Пример данных для замены
    user_replacements = {
        "AL4": "Термометр ТГ-100",
        "AL5": "СР-2024-001",
        "AL6": "2024",
        "AL8": "12",
        "AL9": "15.05.2023",
        "AL10": "0...100",
        "AL11": "0.1",
        "AZ13": "5",
        "AL17": "22.5",
        "A21": "СИ для поверки",
        "AM23": "Да",
        # "J29": 20.1, "L29": 20.2, "N29": 20.0, "P29": 20.3, "R29": 20.1,
        # "AV29": 0.1, "BH29": "±0.5",
        "AM34": "Да",
        "J36": "соответствует",
        "AK38": "Петров А.И.",
        "V41": "10.06.2024"
    }

    # Запуск обработки файла
    replace_template_values(
        file_path="Копия_Термометр_стеклянный_(1).xlsx",
        output_path="Заполненный_протокол.xlsx",
        replacements=user_replacements
    )