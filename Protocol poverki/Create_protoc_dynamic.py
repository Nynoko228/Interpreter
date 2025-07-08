from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.cell import MergedCell

def find_cell_by_text(ws, keyword):
    """Поиск ячейки по ключевому слову"""
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue  # Пропускаем части объединенных ячеек
            if cell.value and keyword.lower() in str(cell.value).lower():
                return cell
    return None

def replace_template_values(file_path, output_path, field_map, replacements):
    wb = load_workbook(filename=file_path)
    ws = wb.active

    # Карта объединённых ячеек
    merged_map = {}
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = (
            merged_range.min_col, merged_range.min_row,
            merged_range.max_col, merged_range.max_row
        )
        top_left = f"{get_column_letter(min_col)}{min_row}"
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                merged_map[f"{get_column_letter(col)}{row}"] = top_left

    # Основной цикл замены
    for key, keyword in field_map.items():
        cell = find_cell_by_text(ws, keyword)
        if not cell:
            print(f"Не найдено поле: {key} (ключевое слово: '{keyword}')")
            continue

        # Обработка объединенных ячеек
        addr = f"{get_column_letter(cell.column)}{cell.row}"
        if addr in merged_map:
            target_addr = merged_map[addr]
            col_letter = ''.join(filter(str.isalpha, target_addr))
            row = int(''.join(filter(str.isdigit, target_addr)))
            col_number = column_index_from_string(col_letter)
        else:
            col_number = cell.column
            row = cell.row

        # Вставка данных (например, в соседнюю ячейку)
        target_cell = ws.cell(row=row, column=col_number + 1)  # +1 колонка справа
        if not isinstance(target_cell, MergedCell):
            value = replacements.get(key)
            if value is not None:
                print(f"Запись в {get_column_letter(col_number + 1)}{row}: {value}")
                target_cell.value = value

    wb.save(output_path)
    print(f"\nФайл успешно сохранён: {output_path}")

if __name__ == "__main__":
    # Карта полей: {ключ: ключевое_слово_в_таблице}
    field_map = {
        "type": "Тип (модель) средства измерений",
        "serial": "Заводской номер",
        "year": "Год выпуска",
        "interval": "Интервал между калибровками",
        "prev_date": "Дата предыдущей калибровки",
        "range": "Пределы измерения",
        "division": "Цена деления",
        "location": "Принадлежность и местонахождение",
        "temp": "Температура окружающего воздуха",
        "result": "1. Внешний осмотр ",
        "conclusion": "Заключение:",
        "executor": "Калибровку выполнил",
        "calibration_date": "Дата калибровки"
    }

    # Данные для замены
    replacements = {
        "type": "Термометр ТГ-100",
        "serial": "СР-2024-001",
        "year": "2024",
        "interval": "12",
        "prev_date": "15.05.2023",
        "range": "0...100",
        "division": "0.1",
        "location": "Цех №5",
        "temp": "22.5",
        "result": "Да",
        "conclusion": "соответствует",
        "executor": "Петров А.И.",
        "calibration_date": "10.06.2024"
    }

    replace_template_values(
        file_path="Копия_Термометр_стеклянный_(1).xlsx",
        output_path="Заполненный_протокол.xlsx",
        field_map=field_map,
        replacements=replacements
    )