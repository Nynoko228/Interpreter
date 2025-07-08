from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

def count_cells_between_markers(file_path, marker1='{{MEASUREMENTS_START}}', marker2='{{MEASUREMENTS_END}}', sheet_name=0):
    """
    Подсчитывает количество ячеек между двумя метками в Excel-файле,
    учитывая объединённые ячейки.

    :param file_path: Путь к Excel-файлу
    :param marker1: Первая метка
    :param marker2: Вторая метка
    :param sheet_name: Имя или индекс листа (по умолчанию первый)
    :return: Словарь с результатами
    """

    def find_marker_in_excel(file_path, marker, sheet_name):
        """Внутренняя функция для поиска метки"""
        wb = load_workbook(filename=file_path)
        sheet = wb[sheet_name] if isinstance(sheet_name, str) else wb.worksheets[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None and str(cell.value) == marker:
                    return (cell.row, cell.column_letter)
        return None

    # Поиск меток
    pos1 = find_marker_in_excel(file_path, marker1, sheet_name)
    pos2 = find_marker_in_excel(file_path, marker2, sheet_name)

    if not pos1 or not pos2:
        print("Не все метки найдены.")
        return None

    row1, col1 = pos1
    row2, col2 = pos2

    # Преобразуем буквы столбцов в индексы
    col1_idx = column_index_from_string(col1)
    col2_idx = column_index_from_string(col2)

    # Определяем границы диапазона
    min_row = min(row1, row2)
    max_row = max(row1, row2)
    min_col = min(col1_idx, col2_idx)
    max_col = max(col1_idx, col2_idx)

    # Переменные для подсчёта
    visible_cells = 0
    merged_ranges_count = 0
    merged_cells_total = 0

    # Загружаем рабочий лист
    wb = load_workbook(filename=file_path)
    sheet = wb[sheet_name] if isinstance(sheet_name, str) else wb.worksheets[sheet_name]

    # Перебираем все ячейки в диапазоне
    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            is_merged = False

            # Проверяем, находится ли ячейка в объединённом диапазоне
            for merged_range in sheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    # Проверяем, является ли ячейка верхней левой в объединённом диапазоне
                    if cell.coordinate == merged_range.start_cell.coordinate:
                        is_merged = True
                        merged_ranges_count += 1
                        visible_cells += 1

                    # Подсчитываем общее количество объединённых ячеек в диапазоне
                    merged_cells_total += (merged_range.max_row - merged_range.min_row + 1) * \
                                          (merged_range.max_col - merged_range.min_col + 1)
                    break

            # Если ячейка не объединена
            if not is_merged:
                visible_cells += 1

    # Возвращаем результат
    return {
        'visible_cells': visible_cells,
        'merged_ranges': merged_ranges_count,
        'merged_cells_total': merged_cells_total
    }

if __name__ == "__main__":
    file_path = 'test.xlsx'
    marker1 = '{{MEASUREMENTS_START}}'
    marker2 = '{{MEASUREMENTS_END}}'

    result = count_cells_between_markers(file_path, marker1, marker2)

    if result:
        print(f"Количество видимых ячеек: {result['visible_cells']}")
        print(f"Количество объединённых диапазонов: {result['merged_ranges']}")
        print(f"Общее количество объединённых ячеек: {result['merged_cells_total']}")
    else:
        print("Не удалось подсчитать ячейки.")