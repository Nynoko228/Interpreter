from openpyxl import load_workbook

def find_marker_in_excel(file_path, marker='{{MEASUREMENTS_START}}', sheet_name=0):
    """
    Ищет заданную метку в Excel-файле, включая объединённые ячейки.

    :param file_path: Путь к файлу Excel
    :param marker: Искомая метка
    :param sheet_name: Имя или индекс листа (по умолчанию первый)
    :return: Кортеж (строка, столбец) или None, если метка не найдена
    """
    try:
        # Загружаем Excel-файл
        wb = load_workbook(filename=file_path)
        sheet = wb.worksheets[sheet_name] if isinstance(sheet_name, int) else wb[sheet_name]

        # Перебираем все ячейки на листе
        for row in sheet.iter_rows():
            for cell in row:
                # Пропускаем пустые ячейки, которые входят в объединённые диапазоны
                if cell.value is not None and str(cell.value) == marker:
                    # Возвращаем позицию в формате Excel (строка и буква столбца)
                    return (cell.row, cell.column_letter)

        return None

    except Exception as e:
        print(f"Ошибка при чтении файла: {e}")
        return None

if __name__ == "__main__":
    file_path = 'test.xlsx'  # Путь к вашему файлу
    marker = '{{MEASUREMENTS_START}}'

    position =  find_marker_in_excel(file_path, marker)

    if position:
        print(f"Метка '{marker}' найдена в ячейке {position[1]}{position[0]}.")
    else:
        print(f"Метка '{marker}' не найдена в файле.")

    marker = '{{MEASUREMENTS_END}}'

    position =  find_marker_in_excel(file_path, marker)

    if position:
        print(f"Метка '{marker}' найдена в ячейке {position[1]}{position[0]}.")
    else:
        print(f"Метка '{marker}' не найдена в файле.")