from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.cell import MergedCell
import re
from collections import defaultdict


class ProtocolGenerator:
    def __init__(self, template_path):
        self.template_path = template_path
        self.placeholders = defaultdict(list)
        self.data_sources = {}
        self.tables = {}

    def add_placeholder(self, name, value):
        """Добавляет значение для простого плейсхолдера"""
        self.placeholders[name].append(value)

    def add_data_source(self, name, data_func):
        """Добавляет источник данных для динамического заполнения"""
        self.data_sources[name] = data_func

    def add_table_config(self, table_name, config):
        """Добавляет конфигурацию таблицы"""
        self.tables[table_name] = config

    def generate(self, output_path):
        """Генерирует протокол на основе шаблона"""
        wb = load_workbook(self.template_path)
        ws = wb.active

        # Замена простых плейсхолдеров
        self._replace_simple_placeholders(ws)

        # Заполнение таблиц
        self._fill_tables(ws)

        # Динамические вычисления
        self._process_data_sources(ws)

        wb.save(output_path)
        return output_path

    def _replace_simple_placeholders(self, ws):
        """Заменяет простые плейсхолдеры вида {{NAME}}"""
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    for ph_name, values in self.placeholders.items():
                        placeholder = f"{{{{{ph_name}}}}}"
                        if placeholder in cell.value:
                            cell.value = cell.value.replace(placeholder, str(values[0]))

    def _fill_tables(self, ws):
        """Заполняет таблицы по конфигурации, учитывая объединённые ячейки"""
        # Строим карту объединённых ячеек
        merged_map = self._build_merged_map(ws)

        for table_name, config in self.tables.items():
            # Поиск начала таблицы по метке
            start_cell = None
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value == f"{{{{{table_name}_START}}}}":
                        start_cell = cell
                        break
                if start_cell:
                    break

            if not start_cell:
                continue

            # Очистка стартовой метки
            self._safe_write(ws, start_cell.coordinate, None, merged_map)

            # Получение данных
            data = self.placeholders.get(table_name, [])

            # Заполнение таблицы
            start_row = start_cell.row + config.get('header_rows', 0)
            start_col = start_cell.column

            for i, row_data in enumerate(data):
                for j, value in enumerate(row_data):
                    cell = ws.cell(
                        row=start_row + i,
                        column=start_col + j
                    )
                    # Безопасная запись с учётом объединённых ячеек
                    self._safe_write(ws, cell.coordinate, value, merged_map)

    def _build_merged_map(self, ws):
        """Создаёт карту объединённых ячеек: {адрес -> главная ячейка}"""
        merged_map = {}
        for merged_range in ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = (
                merged_range.min_col, merged_range.min_row,
                merged_range.max_col, merged_range.max_row
            )
            top_left = f"{get_column_letter(min_col)}{min_row}"
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    addr = f"{get_column_letter(col)}{row}"
                    merged_map[addr] = top_left
        return merged_map

    def _safe_write(self, ws, address, value, merged_map):
        """
        Безопасная запись значения в ячейку с учётом объединённых ячеек
        """
        # Разбор адреса ячейки
        col_letter = ''.join(filter(str.isalpha, address))
        row = int(''.join(filter(str.isdigit, address)))
        col_number = column_index_from_string(col_letter)

        # Проверка, входит ли ячейка в объединённый диапазон
        if address in merged_map:
            target_addr = merged_map[address]
            col_letter = ''.join(filter(str.isalpha, target_addr))
            row = int(''.join(filter(str.isdigit, target_addr)))
            col_number = column_index_from_string(col_letter)

        # Получаем ячейку
        cell = ws.cell(row=row, column=col_number)

        # Проверяем, не является ли ячейка частью объединённой (кроме главной)
        if not isinstance(cell, MergedCell):
            cell.value = value
        else:
            print(f"Пропущено: {address} (часть объединённой ячейки)")

    def _process_data_sources(self, ws):
        """Обрабатывает динамические источники данных"""
        for name, data_func in self.data_sources.items():
            result = data_func()

            # Поиск всех ячеек с этим плейсхолдером
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value == f"{{{{{name}}}}}":
                        cell.value = str(result)


def main():
    # Инициализация генератора
    generator = ProtocolGenerator("Копия_Термометр_стеклянный_(1).xlsx")

    # Добавление статических данных
    generator.add_placeholder("DEVICE_TYPE", "Термометр ТГ-100")
    generator.add_placeholder("SERIAL_NUMBER", "СР-2024-001")
    generator.add_placeholder("MANUFACTURE_YEAR", "2024")

    # Конфигурация таблицы измерений
    generator.add_table_config("MEASUREMENTS", {
        'header_rows': 0,
        'columns': ['T1', 'T2', 'T3', 'T4', 'T5', 'Error', 'Tolerance']
    })

    # Добавление данных измерений
    measurements = [
        # Формат: [Действит. темп., T1, T2, T3, T4, T5, Среднее, Погрешность, Допуск]
        ["20.0°C", 20.1, 20.2, 20.0, 20.3, 20.1, "20.14°C", 0.14, "±0.5"],
        ["40.0°C", 40.0, 40.1, 39.9, 40.2, 40.0, "40.04°C", 0.04, "±0.5"],
        ["60.0°C", 60.2, 60.0, 60.1, 59.9, 60.0, "60.04°C", 0.04, "±0.5"],
        ["80.0°C", 80.1, 80.0, 79.9, 80.2, 80.1, "80.06°C", 0.06, "±0.5"],
        ["100.0°C", 99.9, 100.0, 100.1, 99.8, 100.0, "99.96°C", 0.04, "±0.5"]
    ]
    for row in measurements:
        generator.add_placeholder("MEASUREMENTS", row)

    # Динамические вычисления (пример)
    generator.add_data_source("CONCLUSION",
                              lambda: "соответствует" if all(m[5] < 0.3 for m in measurements) else "не соответствует")

    # Генерация протокола
    output_file = generator.generate("protocol_filled.xlsx")
    print(f"Протокол сгенерирован: {output_file}")


if __name__ == "__main__":
    main()