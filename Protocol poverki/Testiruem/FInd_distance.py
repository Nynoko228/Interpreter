from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, coordinate_to_tuple

def get_merged_ranges_between_markers(
    file_path,
    marker1='{{MEASUREMENTS_START}}',
    marker2='{{MEASUREMENTS_END}}',
    sheet_name=0
):
    """
    Возвращает координаты меток и список объединённых диапазонов между ними,
    отсортированный по строкам (и столбцам при необходимости).

    :param file_path: Путь к Excel-файлу
    :param marker1: Первая метка
    :param marker2: Вторая метка
    :param sheet_name: Имя или индекс листа (по умолчанию первый)
    :return: Словарь с координатами меток и объединёнными диапазонами, или None
    """

    def find_marker(file_path, marker, sheet_name):
        """Внутренняя функция для поиска метки"""
        wb = load_workbook(filename=file_path, read_only=True)
        sheet = wb[sheet_name] if isinstance(sheet_name, str) else wb.worksheets[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None and str(cell.value) == marker:
                    return cell.coordinate  # Например, 'A1'
        return None

    # Поиск меток
    coord1 = find_marker(file_path, marker1, sheet_name)
    coord2 = find_marker(file_path, marker2, sheet_name)

    if not coord1 or not coord2:
        print("Не все метки найдены.")
        return None

    # Парсим координаты меток
    row1, col1 = coordinate_to_tuple(coord1)
    row2, col2 = coordinate_to_tuple(coord2)

    # Определяем границы прямоугольного диапазона
    min_row = min(row1, row2)
    max_row = max(row1, row2)
    min_col = min(col1, col2)
    max_col = max(col1, col2)

    # Загружаем файл и лист
    wb = load_workbook(filename=file_path)
    sheet = wb[sheet_name] if isinstance(sheet_name, str) else wb.worksheets[sheet_name]

    # Список объединённых диапазонов в зоне интереса
    merged_ranges_in_area = []

    # Перебираем все объединённые диапазоны
    for merged_range in sheet.merged_cells.ranges:
        # Получаем границы объединённого диапазона
        mr_min_col, mr_min_row, mr_max_col, mr_max_row = merged_range.bounds

        # Проверяем пересечение с интересующим диапазоном
        if not (
            mr_max_row < min_row or
            mr_min_row > max_row or
            mr_max_col < min_col or
            mr_min_col > max_col
        ):
            # Преобразуем координаты в формат A1:B2
            start_col = get_column_letter(mr_min_col)
            end_col = get_column_letter(mr_max_col)
            start = f"{start_col}{mr_min_row}"
            end = f"{end_col}{mr_max_row}"
            merged_ranges_in_area.append({
                'range': f"{start}:{end}",
                'min_row': mr_min_row,
                'min_col': mr_min_col
            })

    # Сортируем диапазоны по строке, затем по столбцу
    merged_ranges_in_area.sort(key=lambda x: (x['min_row'], x['min_col']))

    # Возвращаем только строки диапазонов
    sorted_ranges = [item['range'] for item in merged_ranges_in_area]

    # Возвращаем результат
    return {
        "marker1": coord1,
        "marker2": coord2,
        "merged_ranges": sorted_ranges
    }

if __name__ == "__main__":
    file_path = 'test.xlsx'
    marker1 = '{{MEASUREMENTS_START}}'
    marker2 = '{{MEASUREMENTS_END}}'

    result = get_merged_ranges_between_markers(file_path, marker1, marker2)

    if result:
        print(f"Метка 1: {result['marker1']}")
        print(f"Метка 2: {result['marker2']}")
        print("Объединённые диапазоны между метками:")
        for merged_range in result['merged_ranges']:
            print(f" - {merged_range}")
    else:
        print("Не удалось найти метки или объединённые диапазоны.")