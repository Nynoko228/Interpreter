import openpyxl
import re

def find_all_tags(file_path):
    wb = openpyxl.load_workbook(file_path)
    results = []

    # Регулярное выражение для поиска всех меток вида {{...}}
    pattern = re.compile(r'\{\{(.*?)\}\}')

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Множество координат, входящих в объединённые ячейки, но не являющихся первыми
        merged_cells_set = set()
        # Заполняем множество пропускаемых ячеек
        for merged_range in ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    if row != min_row or col != min_col:
                        merged_cells_set.add((row, col))
        # Перебираем все ячейки
        for row in ws.iter_rows():
            for cell in row:
                # Пропускаем ячейки, не являющиеся первыми в объединённых диапазонах
                if (cell.row, cell.column) in merged_cells_set:
                    continue

                # Проверяем, есть ли значение и является ли оно строкой
                if cell.value and isinstance(cell.value, str):
                    matches = pattern.findall(cell.value)
                    for tag in matches:
                        results.append((sheet_name, cell.row, cell.column, tag.strip()))

    return results


# core/protocol_utils.py
import openpyxl
import re


def find_all_tags1(file_path):
    wb = openpyxl.load_workbook(file_path)
    results = {}

    # Регулярное выражение для поиска меток вида {{...}}
    pattern = re.compile(r'\{\{(.*?)\}\}')

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_tags = {}

        # Собираем объединенные ячейки для исключения дубликатов
        merged_cells = set()
        for merge_range in ws.merged_cells.ranges:
            min_row, min_col, max_row, max_col = merge_range.min_row, merge_range.min_col, merge_range.max_row, merge_range.max_col
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    if row != min_row or col != min_col:
                        merged_cells.add((row, col))

        # Обрабатываем все ячейки в листе
        for row in ws.iter_rows():
            for cell in row:
                # Пропускаем ячейки в объединенных областях
                if (cell.row, cell.column) in merged_cells:
                    continue

                # Ищем метки в значении ячейки
                if cell.value and isinstance(cell.value, str):
                    matches = pattern.findall(cell.value)
                    for tag in matches:
                        clean_tag = tag.strip()
                        sheet_tags[clean_tag] = [cell.row, cell.column]

        if sheet_tags:
            results[sheet_name] = sheet_tags

    return results

# Пример использования
if __name__ == "__main__":
    file_path = 'Копия_Термометр_стеклянный_(1).xlsx'
    found_tags = find_all_tags1(file_path)
    print("Найденные метки:")
    print(found_tags)
    # all_tags = {}
    # for sheet, row, col, tag in found_tags:
    #     if sheet not in all_tags:
    #         all_tags[sheet] = {tag: [row, col]}
    #     else:
    #         if tag not in all_tags[sheet]:
    #             all_tags[sheet].update({tag: [row, col]})
    #         else:
    #             all_tags[sheet][tag] = [row, col]
    # print(all_tags)